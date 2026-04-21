from __future__ import annotations

import argparse
import math
import re
import time
import unicodedata
import json
import os
import requests
from collections import Counter
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

DATA_PATH = Path(__file__).with_name("tmdb_top1000_movies.xlsx")
DESCRIPTION_LIMIT = 500
MODEL = "gemma4:31b-cloud"
OLLAMA_HOST = "https://ollama.com"
LLM_CHAR_BUDGET = 420
REQUEST_TIMEOUT_SECONDS = 8

STOPWORDS = {
    "a", "about", "after", "all", "an", "and", "any", "are", "as", "at", "be",
    "because", "but", "by", "for", "from", "have", "i", "if", "im", "in", "into",
    "is", "it", "its", "just", "kind", "like", "me", "movie", "movies", "my",
    "of", "on", "or", "something", "that", "the", "them", "these", "those", "to",
    "want", "with",
}

GENRE_ALIASES = {
    "action": {"action", "adrenaline", "explosive", "fight", "fight scenes"},
    "adventure": {"adventure", "quest", "journey", "epic"},
    "animation": {"animated", "animation", "cartoon", "pixar", "disney"},
    "comedy": {"comedy", "comedic", "funny", "humor", "laughs", "witty"},
    "crime": {"crime", "gangster", "heist", "mafia", "mob"},
    "drama": {"drama", "dramatic", "character study", "emotional"},
    "family": {"family", "kids", "kid", "all ages", "family night"},
    "fantasy": {"fantasy", "magic", "magical", "mythic"},
    "history": {"historical", "history", "period piece"},
    "horror": {"horror", "scary", "creepy", "terrifying", "haunted"},
    "music": {"music", "musical", "songs", "singing"},
    "mystery": {"mystery", "whodunit", "investigation", "twist"},
    "romance": {"romance", "romantic", "rom-com", "romcom", "love story"},
    "science fiction": {"sci-fi", "science fiction", "space", "futuristic"},
    "superhero": {"superhero", "superheroes", "marvel", "dc comics", "comic book"},
    "thriller": {"thriller", "tense", "suspense", "suspenseful"},
    "war": {"war", "military", "battlefront"},
}

PHRASE_HINTS = {
    "superhero": ["superhero", "comic", "hero", "villain", "marvel", "dc"],
    "buddy cop": ["buddy", "banter", "team-up", "partners", "crime", "cop"],
    "feel good": ["uplifting", "warm", "fun", "heart", "hopeful"],
    "emotional": ["moving", "heartwarming", "character", "drama"],
    "heart": ["heartwarming", "warm", "hopeful"],
    "twist ending": ["twist", "reveal", "mystery", "psychological", "mind-bending"],
    "slow burn": ["atmospheric", "patient", "moody", "character", "tension"],
    "exciting": ["action", "thriller", "intense", "momentum"],
    "not mindless": ["smart", "clever", "character", "driven"],
    "dread": ["psychological", "atmospheric", "tension"],
    "chemistry": ["romance", "spark", "relationship"],
    "unreliable": ["psychological", "paranoia", "mystery"],
    "family movie night": ["family", "heartwarming", "adventure", "fun", "crowd-pleasing"],
    "rom com": ["romance", "chemistry", "charming", "comedy"],
    "mind bending": ["mind-bending", "dream", "reality", "psychological", "mystery"],
}

NEGATION_PHRASES = [
    "no ", "not ", "avoid ", "hate ", "dislike ", "dont want ", "don't want ", "without ",
]


@dataclass(frozen=True)
class Movie:
    tmdb_id: int
    title: str
    year: int | None
    runtime_min: int | None
    genres: tuple[str, ...]
    overview: str
    tagline: str
    director: str
    cast: tuple[str, ...]
    keywords: tuple[str, ...]
    popularity: float
    vote_average: float
    vote_count: int
    us_rating: str
    normalized_title: str
    title_key: str
    token_set: frozenset[str]
    searchable_blob: str
    quality_score: float


def _ascii_text(text: str) -> str:
    return unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode("ascii")


def normalize_text(text: str) -> str:
    lowered = _ascii_text(text).lower()
    lowered = re.sub(r"&", " and ", lowered)
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def title_key(text: str) -> str:
    cleaned = normalize_text(text)
    cleaned = re.sub(r"\b(the|a|an)\b\s*", "", cleaned)
    cleaned = re.sub(r"\b\d{4}\b", "", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def tokenize(text: str) -> list[str]:
    return [token for token in normalize_text(text).split() if len(token) > 2 and token not in STOPWORDS]


def split_csvish(text: str | None) -> tuple[str, ...]:
    if not text:
        return ()
    return tuple(part.strip() for part in str(text).split(",") if part and part.strip())


def _safe_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _safe_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


@lru_cache(maxsize=1)
def load_movies() -> tuple[Movie, ...]:
    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    index = {name: idx for idx, name in enumerate(headers)}

    raw_movies: list[dict[str, object]] = []
    max_popularity = 1.0
    max_vote = 1.0
    max_vote_count = 1.0

    for row in rows:
        tmdb_id = _safe_int(row[index["tmdb_id"]])
        title = str(row[index["title"]] or "").strip()
        if not tmdb_id or not title:
            continue

        popularity = _safe_float(row[index["popularity"]])
        vote_average = _safe_float(row[index["vote_average"]])
        vote_count = _safe_int(row[index["vote_count"]]) or 0

        raw_movie = {
            "tmdb_id": tmdb_id,
            "title": title,
            "year": _safe_int(row[index["year"]]),
            "runtime_min": _safe_int(row[index["runtime_min"]]),
            "genres": split_csvish(row[index["genres"]]),
            "overview": str(row[index["overview"]] or "").strip(),
            "tagline": str(row[index["tagline"]] or "").strip(),
            "director": str(row[index["director"]] or "").strip(),
            "cast": split_csvish(row[index["top_cast"]]),
            "keywords": split_csvish(row[index["keywords"]]),
            "popularity": popularity,
            "vote_average": vote_average,
            "vote_count": vote_count,
            "us_rating": str(row[index["us_rating"]] or "").strip(),
        }

        raw_movies.append(raw_movie)
        max_popularity = max(max_popularity, popularity)
        max_vote = max(max_vote, vote_average)
        max_vote_count = max(max_vote_count, float(vote_count))

    movies: list[Movie] = []
    for raw_movie in raw_movies:
        joined = " ".join(
            [
                str(raw_movie["title"]),
                " ".join(raw_movie["genres"]),
                str(raw_movie["overview"]),
                str(raw_movie["tagline"]),
                str(raw_movie["director"]),
                " ".join(raw_movie["cast"]),
                " ".join(raw_movie["keywords"]),
            ]
        )
        tokens = frozenset(tokenize(joined))
        popularity = float(raw_movie["popularity"])
        vote_average = float(raw_movie["vote_average"])
        vote_count = int(raw_movie["vote_count"])

        quality_score = (
            0.5 * (vote_average / max_vote)
            + 0.2 * math.log1p(popularity) / math.log1p(max_popularity)
            + 0.3 * math.log1p(vote_count) / math.log1p(max_vote_count)
        )

        movies.append(
            Movie(
                tmdb_id=int(raw_movie["tmdb_id"]),
                title=str(raw_movie["title"]),
                year=raw_movie["year"],
                runtime_min=raw_movie["runtime_min"],
                genres=tuple(raw_movie["genres"]),
                overview=str(raw_movie["overview"]),
                tagline=str(raw_movie["tagline"]),
                director=str(raw_movie["director"]),
                cast=tuple(raw_movie["cast"]),
                keywords=tuple(raw_movie["keywords"]),
                popularity=popularity,
                vote_average=vote_average,
                vote_count=vote_count,
                us_rating=str(raw_movie["us_rating"]),
                normalized_title=normalize_text(str(raw_movie["title"])),
                title_key=title_key(str(raw_movie["title"])),
                token_set=tokens,
                searchable_blob=normalize_text(joined),
                quality_score=quality_score,
            )
        )

    return tuple(movies)


@lru_cache(maxsize=1)
def movie_lookup() -> dict[int, Movie]:
    return {movie.tmdb_id: movie for movie in load_movies()}


@lru_cache(maxsize=1)
def title_lookup() -> dict[str, list[Movie]]:
    mapping: dict[str, list[Movie]] = {}
    for movie in load_movies():
        mapping.setdefault(movie.title_key, []).append(movie)
    return mapping


@lru_cache(maxsize=1)
def token_idf() -> dict[str, float]:
    movies = load_movies()
    doc_count = len(movies)
    counts: Counter[str] = Counter()
    for movie in movies:
        counts.update(movie.token_set)
    return {token: math.log((doc_count + 1) / (count + 1)) + 1.0 for token, count in counts.items()}


def heuristic_extract_preferences(preferences: str) -> dict[str, object]:
    normalized = normalize_text(preferences)
    tokens = tokenize(preferences)
    token_weights: Counter[str] = Counter()
    preferred_genres: set[str] = set()
    avoided_genres: set[str] = set()
    explicit_exclusions: set[str] = set()

    for token in tokens:
        token_weights[token] += 2.0

    for phrase, hints in PHRASE_HINTS.items():
        phrase_norm = normalize_text(phrase)
        if phrase_norm in normalized:
            is_negated = any(neg + phrase_norm in normalized for neg in NEGATION_PHRASES)
            boost = -3.0 if is_negated else 2.5
            for hint in hints:
                for token in tokenize(hint):
                    token_weights[token] += boost

    for genre, aliases in GENRE_ALIASES.items():
        for alias in aliases | {genre}:
            alias_norm = normalize_text(alias)
            if alias_norm and alias_norm in normalized:
                is_negated = False
                for neg in NEGATION_PHRASES:
                    idx = normalized.find(neg)
                    while idx >= 0:
                        window = normalized[idx: idx + 80]
                        if alias_norm in window:
                            is_negated = True
                            break
                        idx = normalized.find(neg, idx + 1)
                    if is_negated:
                        break
                if is_negated:
                    avoided_genres.add(genre)
                else:
                    preferred_genres.add(genre)
                    for token in tokenize(alias):
                        token_weights[token] += 3.5

    for negation in NEGATION_PHRASES:
        start = 0
        while True:
            idx = normalized.find(negation, start)
            if idx < 0:
                break
            fragment = normalized[idx + len(negation) : idx + len(negation) + 60]
            for token in fragment.split()[:5]:
                if len(token) > 2 and token not in STOPWORDS:
                    explicit_exclusions.add(token)
            start = idx + len(negation)

    return {
        "normalized": normalized,
        "token_weights": token_weights,
        "preferred_genres": preferred_genres,
        "avoided_genres": avoided_genres,
        "excluded_tokens": explicit_exclusions,
    }


def agentic_extract_intent(preferences: str) -> dict | None:
    api_key = os.getenv("OLLAMA_API_KEY")
    if not api_key:
        return None

    prompt = f"""
Analyze this movie request: "{preferences}"

Extract the user's intent into exactly this JSON format. No markdown or text.
{{
  "must_have_genres": [],
  "must_not_have_genres": [],
  "target_era": "any", 
  "discovery_mode": "neutral", 
  "vibe_keywords": []
}}

Rules for values:
target_era choices: "classic" (pre-1990), "90s" (1990-1999), "2000s", "recent" (2015+), "any"
discovery_mode choices: "hidden_gem" (unknown/underrated), "blockbuster" (popular/famous), "neutral"
vibe_keywords: max 5 english words describing pacing or vibe (e.g. short, fast, creepy)
"""

    try:
        response = requests.post(
            f"{OLLAMA_HOST}/api/chat",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "model": MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are an intent extractor. Return ONLY valid JSON."
                    },
                    {"role": "user", "content": prompt},
                ],
                "stream": False,
            },
            timeout=3,
        )
        response.raise_for_status()
        content = response.json().get("message", {}).get("content", "").strip()
        match = re.search(r"\{.*\}", content, re.DOTALL)
        if match:
            parsed = json.loads(match.group(0))
            return parsed
    except Exception:
        pass
    return None


def extract_preferences(preferences: str) -> dict[str, object]:
    signals = heuristic_extract_preferences(preferences)
    agent_signals = agentic_extract_intent(preferences)
    if agent_signals:
        signals["agent_intent"] = agent_signals
    return signals


def watched_movie_ids(history: list[str], history_ids: list[int]) -> set[int]:
    watched_ids = {int(movie_id) for movie_id in history_ids if isinstance(movie_id, int) or str(movie_id).isdigit()}
    for title in history:
        key = title_key(title)
        if key and key in title_lookup():
            watched_ids.update(movie.tmdb_id for movie in title_lookup()[key])
    return watched_ids


def build_history_profile(history: list[str], history_ids: list[int]) -> dict[str, Counter[str]]:
    lookup = movie_lookup()
    profile = {
        "genres": Counter(),
        "keywords": Counter(),
        "cast": Counter(),
        "director": Counter(),
        "tokens": Counter(),
    }

    for movie_id in watched_movie_ids(history, history_ids):
        movie = lookup.get(movie_id)
        if not movie:
            continue
        profile["genres"].update(normalize_text(genre) for genre in movie.genres)
        profile["keywords"].update(tokenize(" ".join(movie.keywords)))
        profile["cast"].update(tokenize(" ".join(movie.cast[:3])))
        profile["director"].update(tokenize(movie.director))
        profile["tokens"].update(movie.token_set)

    return profile


def build_query_weights(
    preferences: str, history: list[str], history_ids: list[int]
) -> tuple[Counter[str], dict[str, object], dict[str, Counter[str]]]:
    signals = extract_preferences(preferences)
    history_profile = build_history_profile(history, history_ids)
    query_weights: Counter[str] = Counter(signals["token_weights"])

    if "agent_intent" in signals:
        vibe_keywords = signals["agent_intent"].get("vibe_keywords", [])
        for vibe in vibe_keywords:
            for token in tokenize(vibe):
                query_weights[token] += 3.0

    history_strength = 1.0 if len(query_weights) < 4 and not signals["preferred_genres"] else 0.35

    for genre, count in history_profile["genres"].most_common(4):
        for token in tokenize(genre):
            query_weights[token] += history_strength * min(count, 2)

    for token, count in history_profile["keywords"].most_common(10):
        query_weights[token] += 0.35 * min(count, 3)

    for token, count in history_profile["cast"].most_common(6):
        query_weights[token] += 0.25 * min(count, 2)

    for token, count in history_profile["director"].most_common(3):
        query_weights[token] += 0.6 * min(count, 2)

    return query_weights, signals, history_profile


def score_movie(
    movie: Movie,
    query_weights: Counter[str],
    signals: dict[str, object],
    history_profile: dict[str, Counter[str]],
    watched_ids: set[int],
) -> float:
    if movie.tmdb_id in watched_ids:
        return float("-inf")

    idf = token_idf()
    score = 0.0

    for token, weight in query_weights.items():
        if token in movie.token_set:
            score += weight * idf.get(token, 1.0)

    movie_genres = {normalize_text(genre) for genre in movie.genres}

    for genre in signals["preferred_genres"]:
        if genre in movie_genres:
            score += 8.0
        else:
            score -= 6.0

    for genre in signals["avoided_genres"]:
        if genre in movie_genres:
            score -= 25.0

    matched_preferred = sum(1 for genre in signals["preferred_genres"] if genre in movie_genres)
    preferred_count = len(signals["preferred_genres"])
    if preferred_count >= 2 and matched_preferred == preferred_count:
        score += 6.0
    elif preferred_count >= 2 and matched_preferred == 0:
        score -= 6.0

    for token in signals["excluded_tokens"]:
        if token in movie.token_set or token in movie.searchable_blob:
            score -= 5.0

    shared_history_genres = sum(history_profile["genres"].get(genre, 0) for genre in movie_genres)
    score += min(shared_history_genres, 4) * 0.9

    keyword_overlap = sum(history_profile["keywords"].get(token, 0) for token in movie.token_set)
    score += min(keyword_overlap, 10) * 0.18

    if movie.director:
        director_overlap = sum(history_profile["director"].get(token, 0) for token in tokenize(movie.director))
        score += director_overlap * 0.8

    cast_overlap = sum(history_profile["cast"].get(token, 0) for token in tokenize(" ".join(movie.cast[:3])))
    score += min(cast_overlap, 3) * 0.5

    agent_intent = signals.get("agent_intent", {})
    if agent_intent:
        must_have = {normalize_text(g) for g in agent_intent.get("must_have_genres", [])}
        must_have.update(signals["preferred_genres"])
        
        must_not_have = {normalize_text(g) for g in agent_intent.get("must_not_have_genres", [])}
        must_not_have.update(signals["avoided_genres"])
        
        for g in must_have:
            if g and g not in movie_genres:
                score -= 30.0
        for g in must_not_have:
            if g and g in movie_genres:
                score -= 30.0

        target_era = agent_intent.get("target_era", "any")
        if target_era != "any" and movie.year:
            if target_era == "classic" and movie.year < 1990: score += 6.0
            elif target_era == "90s" and 1990 <= movie.year <= 1999: score += 6.0
            elif target_era == "2000s" and 2000 <= movie.year <= 2014: score += 6.0
            elif target_era == "recent" and movie.year >= 2015: score += 6.0
            else: score -= 3.0

        discovery_mode = agent_intent.get("discovery_mode", "neutral")
        if discovery_mode == "hidden_gem":
            score += (movie.vote_average - 7.0) * 4.0
            score -= min(movie.popularity / 50.0, 8.0) 
            score += movie.quality_score * 0.5 
        elif discovery_mode == "blockbuster":
            score += math.log1p(movie.popularity) * 2.0
            score += movie.quality_score * 3.0
        else:
            score += movie.quality_score * 3.0
    else:
        score += movie.quality_score * 3.0

    if movie.vote_average < 6.0:
        score -= 6.0
    elif movie.vote_average < 6.5:
        score -= 2.5

    if movie.vote_average >= 7.5:
        score += 1.5

    if movie.vote_count >= 5000:
        score += 1.2

    if movie.us_rating == "R" and "family" in signals["normalized"]:
        score -= 8.0

    return score


def explain_match(
    movie: Movie, preferences: str, signals: dict[str, object], history_profile: dict[str, Counter[str]]
) -> list[str]:
    reasons: list[str] = []
    movie_genres = {normalize_text(value) for value in movie.genres}

    preferred_genres = [genre for genre in signals["preferred_genres"] if genre in movie_genres]
    if preferred_genres:
        reasons.append(f"leans into your taste for {', '.join(preferred_genres[:2])}")

    matched_keywords = [token for token, _ in signals["token_weights"].most_common(12) if token in movie.token_set]
    if matched_keywords:
        reasons.append(f"matches the {', '.join(matched_keywords[:3])} vibe you're asking for")

    history_genre_matches = [
        genre for genre, count in history_profile["genres"].most_common() if count and genre in movie_genres
    ]
    if history_genre_matches:
        reasons.append("fits the patterns in your watch history")

    if movie.vote_average >= 7.8:
        reasons.append("has the kind of quality that usually makes a pick feel worth it")

    return reasons[:3]


def enforce_description_limit(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if len(cleaned) <= DESCRIPTION_LIMIT:
        return cleaned
    truncated = cleaned[: DESCRIPTION_LIMIT - 1].rstrip()
    if ". " in truncated:
        truncated = truncated.rsplit(". ", 1)[0].rstrip(". ")
    return truncated[: DESCRIPTION_LIMIT - 1].rstrip() + "..."


def deterministic_description(movie: Movie, preferences: str, history: list[str], reasons: list[str], start_time: float) -> str:
    movie_genres = {normalize_text(g) for g in movie.genres}
    pref_text = normalize_text(preferences)
    keyword_text = " ".join(movie.keywords).lower()

    if "animation" in movie_genres and any(word in pref_text for word in ["emotional", "heart", "animated", "warm"]):
        hook = f"If you want something animated that feels heartfelt without being overly sweet, {movie.title} is a great call."
    elif "romance" in movie_genres and "comedy" in movie_genres:
        hook = f"If you want something charming, funny, and easy to fall into, {movie.title} is a great call."
    elif "horror" in movie_genres:
        hook = f"If you want something genuinely creepy that still feels like a smart pick, {movie.title} looks especially well chosen."
    elif "thriller" in movie_genres or "mystery" in movie_genres:
        hook = f"If you're in the mood for something tense, sharp, and satisfying, {movie.title} is an easy one to get excited about."
    elif "science fiction" in movie_genres:
        hook = f"If you want something imaginative with real momentum, {movie.title} feels like a very fun choice."
    else:
        hook = f"If you want something that really fits this mood, {movie.title} is a strong choice."

    fit_reasons = []

    if "animation" in movie_genres:
        fit_reasons.append("warmth")
    if "drama" in movie_genres:
        fit_reasons.append("emotional insight")
    if "comedy" in movie_genres:
        fit_reasons.append("personality")
    if "romance" in movie_genres:
        fit_reasons.append("chemistry")
    if "thriller" in movie_genres:
        fit_reasons.append("tension")
    if "mystery" in movie_genres:
        fit_reasons.append("intrigue")
    if "science fiction" in movie_genres:
        fit_reasons.append("an imaginative style")
    if "adventure" in movie_genres:
        fit_reasons.append("forward momentum")
    if "family" in movie_genres:
        fit_reasons.append("heart")

    if "music" in keyword_text or "jazz" in keyword_text:
        fit_reasons.append("a fresh creative energy")
    elif "friendship" in keyword_text or "family" in keyword_text:
        fit_reasons.append("a touching emotional core")
    elif "dream" in keyword_text or "reality" in keyword_text:
        fit_reasons.append("a memorable imaginative hook")
    elif movie.tagline:
        fit_reasons.append("a memorable hook")

    for reason in reasons:
        short_reason = reason.replace("leans into your taste for ", "").replace("matches the ", "")
        if short_reason and len(fit_reasons) < 4:
            fit_reasons.append(short_reason)

    deduped = []
    seen = set()
    for reason in fit_reasons:
        if reason not in seen:
            deduped.append(reason)
            seen.add(reason)

    fit_reasons = deduped[:3]

    if len(fit_reasons) >= 3:
        fit_sentence = (
            f"It blends {fit_reasons[0]}, {fit_reasons[1]}, and {fit_reasons[2]} "
            "in a way that feels vivid rather than generic."
        )
    elif len(fit_reasons) == 2:
        fit_sentence = f"It blends {fit_reasons[0]} and {fit_reasons[1]} in a way that feels vivid rather than generic."
    elif len(fit_reasons) == 1:
        fit_sentence = f"It leans into {fit_reasons[0]} in a way that feels specific and memorable."
    else:
        fit_sentence = "It has the kind of tone and payoff that makes it feel like a genuinely good pick."

    if "animation" in movie_genres and "drama" in movie_genres:
        payoff = "It's the kind of movie that leaves you moved, lighter, and glad you picked it."
    elif "romance" in movie_genres and "comedy" in movie_genres:
        payoff = "It feels breezy, appealing, and fun to spend time with from the start."
    elif "thriller" in movie_genres or "mystery" in movie_genres:
        payoff = "It has the kind of payoff that makes the choice feel satisfying instead of obvious."
    elif "horror" in movie_genres:
        payoff = "It feels intense in the right way and sticks with you after it ends."
    elif "science fiction" in movie_genres:
        payoff = "It feels expansive and entertaining without losing the human side of the story."
    else:
        payoff = "It feels like the kind of pick that lands well once you actually press play."

    description = f"{hook} {fit_sentence} {payoff}"
    return enforce_description_limit(description)


def agentic_judge_and_describe(movies: list[Movie], preferences: str, history: list[str], signals: dict[str, object], history_profile: dict[str, Counter[str]], start_time: float) -> dict:
    fallback_movie = movies[0]
    fallback_reasons = explain_match(fallback_movie, preferences, signals, history_profile)
    fallback_desc = deterministic_description(fallback_movie, preferences, history, fallback_reasons, start_time)
    fallback_candidate = {"tmdb_id": fallback_movie.tmdb_id, "description": fallback_desc}

    api_key = os.getenv("OLLAMA_API_KEY")
    remaining_budget = REQUEST_TIMEOUT_SECONDS - (time.monotonic() - start_time)
    if not api_key or remaining_budget < 2.0:
        return fallback_candidate

    candidates_text = ""
    for i, m in enumerate(movies):
        candidates_text += f"[{i+1}] {m.title} ({m.year or 'Unknown'}) | Runtime: {m.runtime_min or 'Unknown'} min | tmdb_id: {m.tmdb_id}\nGenres: {', '.join(m.genres)}\nOverview: {m.overview}\n\n"

    prompt = (
        "You are an expert, emotionally intelligent movie recommendation agent.\n"
        f"User preferences: {preferences}\n"
        f"Watch history: {', '.join(history[:8]) if history else 'None provided'}\n\n"
        "Here are the top candidates that match their taste profile:\n"
        f"{candidates_text}"
        "Task:\n"
        "1. Act as a judge. Pick the single best candidate that fits the user's preferences.\n"
        f"2. Write a persuasive, emotionally resonant recommendation blurb (max {LLM_CHAR_BUDGET} chars, no spoilers, no bullet points). "
        "Follow these 5 rules for the blurb:\n"
        "   RULE 1 - OPEN WITH HISTORY: Start by referencing a specific movie from their Watch history that shares DNA with your pick (e.g. 'Since you loved [watched movie]...'). Skip if no history.\n"
        "   RULE 2 - WEAVE IN RUNTIME: In the first two sentences, naturally embed the runtime (e.g. 'In this gripping [X]-minute thriller...'). Do NOT skip this if runtime is available.\n"
        "   RULE 3 - VIVID DESCRIPTION: Use 1-2 sentences of vivid, specific, emotionally charged language about the experience. Avoid generic words like 'great' or 'amazing'.\n"
        "   RULE 4 - DEFEND NEGATIVES: If the user expressed hates or avoidances, actively rebut them in your blurb (e.g. 'This isn't a loud action flick, but a...'). This builds trust.\n"
        "   RULE 5 - COMPELLING CLOSE: End with one sentence that makes them want to press play right now.\n"
        "3. Output ONLY a valid JSON object matching this exact shape:\n"
        '{"thought_process": "<why this movie perfectly matches in 15 words>", "tmdb_id": <selected tmdb_id integer>, "description": "<your blurb here>"}\n'
    )

    try:
        response = requests.post(
            f"{OLLAMA_HOST}/api/chat",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "model": MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are a movie recommendation AI. You MUST output ONLY valid JSON without markdown wrapping."
                    },
                    {"role": "user", "content": prompt},
                ],
                "stream": False,
            },
            timeout=(2, max(2.0, min(remaining_budget, REQUEST_TIMEOUT_SECONDS))),
        )
        response.raise_for_status()
        content = response.json().get("message", {}).get("content", "").strip()

        match = re.search(r"\{.*\}", content, re.DOTALL)
        if match:
            parsed = json.loads(match.group(0))
            if "tmdb_id" in parsed and "description" in parsed:
                return {
                    "tmdb_id": int(parsed["tmdb_id"]),
                    "description": enforce_description_limit(str(parsed["description"]))
                }
        raise ValueError("Invalid JSON schema returned by LLM.")
    except Exception:
        return fallback_candidate


def validate_output(candidate: dict[str, object], watched_ids: set[int]) -> dict[str, object]:
    movie_id = candidate.get("tmdb_id")
    description = str(candidate.get("description", "")).strip()

    if not isinstance(movie_id, int):
        raise ValueError("tmdb_id must be an int")
    if movie_id not in movie_lookup():
        raise ValueError("tmdb_id not in candidate list")
    if movie_id in watched_ids:
        raise ValueError("recommended movie is already watched")
    if not description:
        raise ValueError("description cannot be empty")

    movie = movie_lookup()[movie_id]

    return {
        "tmdb_id": movie_id,
        "movie_info": {
            "title": movie.title,
            "year": movie.year,
            "runtime_min": movie.runtime_min,
            "director": movie.director,
            "genres": list(movie.genres),
            "vote_average": movie.vote_average,
        },
        "description": enforce_description_limit(description),
    }


def choose_top_movies(
    preferences: str, history: list[str], history_ids: list[int], top_k: int = 5
) -> tuple[list[Movie], dict[str, object], dict[str, Counter[str]], set[int]]:
    watched_ids = watched_movie_ids(history, history_ids)
    query_weights, signals, history_profile = build_query_weights(preferences, history, history_ids)

    scored = []
    for movie in load_movies():
        score = score_movie(movie, query_weights, signals, history_profile, watched_ids)
        if math.isfinite(score):
            scored.append((score, movie))

    scored.sort(key=lambda item: item[0], reverse=True)
    top_movies = [movie for score, movie in scored[:top_k]]

    if not top_movies:
        fallback = max(
            (movie for movie in load_movies() if movie.tmdb_id not in watched_ids),
            key=lambda item: item.quality_score,
        )
        top_movies = [fallback]

    return top_movies, signals, history_profile, watched_ids


def get_recommendation(preferences: str, history: list[str], history_ids: list[int] = []) -> dict:
    start_time = time.monotonic()
    movies, signals, history_profile, watched_ids = choose_top_movies(preferences, history, history_ids, top_k=5)
    candidate = agentic_judge_and_describe(movies, preferences, history, signals, history_profile, start_time)
    return validate_output(candidate, watched_ids)


def _parse_history_arg(values: Iterable[str] | None) -> list[str]:
    if not values:
        return []
    items: list[str] = []
    for value in values:
        for part in str(value).split("|"):
            cleaned = part.strip()
            if cleaned:
                items.append(cleaned)
    return items


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a movie recommendation from the BAMS 521 candidate set.")
    parser.add_argument("--preferences", help="Free-text user preferences.")
    parser.add_argument("--history", nargs="*", help="Watched movie titles. Use repeated args or pipe-separated values.")
    parser.add_argument("--history-ids", nargs="*", type=int, help="TMDB IDs corresponding to the watch history.")
    args = parser.parse_args()

    preferences = args.preferences or input("Preferences: ").strip()
    history = (
        _parse_history_arg(args.history)
        if args.history is not None
        else _parse_history_arg([input("Watch history (optional): ").strip()])
    )
    history_ids = args.history_ids or []
    print(get_recommendation(preferences, history, history_ids))


if __name__ == "__main__":
    main()
